import random
from django.core.mail import send_mail
from openpyxl import load_workbook
from django.utils import timezone
from myapp.models import Membresia
import requests

def generate_code():
    return str(random.randint(100000, 999999))

def send_verification_email(email, code):
    subject = 'Código de Verificación'
    message = f'Tu código de verificación es: {code}'
    from_email = 'tu-correo@gmail.com'  # O el configurado en settings.py
    recipient_list = [email]

    send_mail(subject, message, from_email, recipient_list)






def leer_excel_con_estilos(path):
    wb = load_workbook(path)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    data = []
    styles = []

    for row in sheet.iter_rows(min_row=2, values_only=False):
        row_data = []
        row_styles = []

        for cell in row:
            row_data.append(cell.value)
            fill = cell.fill
            color = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == 'rgb' else None
            row_styles.append({'color': f'#{color[-6:]}' if color else None})

        data.append(row_data)
        styles.append(row_styles)

    return headers, data, styles


def enviar_notificaciones_membresia():
    hoy = timezone.now().date()
    membresias = Membresia.objects.filter(pagado='si')

    for m in membresias:
        if m.fecha_expiracion:
            dias_restantes = (m.fecha_expiracion - hoy).days

            if dias_restantes <= 0:
                asunto = 'Tu membresía ha vencido'
                mensaje = (
                    f'Hola {m.usuario.first_name},\n\n'
                    'Tu membresía ha vencido. Por favor, realiza el pago para continuar disfrutando del servicio.'
                )
                send_mail(
                    asunto,
                    mensaje,
                    None,  # Usará DEFAULT_FROM_EMAIL
                    [m.usuario.email],
                    fail_silently=False,
                )

            elif dias_restantes <= 2:
                asunto = 'Tu membresía está por vencer'
                mensaje = (
                    f'Hola {m.usuario.first_name},\n\n'
                    f'Te recordamos que tu membresía vence en {dias_restantes} día(s). '
                    'No olvides renovar para evitar interrupciones en el servicio.'
                )
                send_mail(
                    asunto,
                    mensaje,
                    None,
                    [m.usuario.email],
                    fail_silently=False,
                )



def get_geo_from_ip(ip):
    try:
        response = requests.get(f'https://ipapi.co/{ip}/json/')
        data = response.json()
        ciudad = data.get('city', 'Desconocido')
        region = data.get('region', 'Desconocido')
    except Exception:
        ciudad = 'Desconocido'
        region = 'Desconocido'
    return {'ciudad': ciudad, 'region': region}
